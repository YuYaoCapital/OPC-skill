#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
R9Alpha 研基基金评价报告生成器 V2

用法：
    python generate-fund-evaluation.py <基金代码> [--template <模板路径>] [--output-dir <输出目录>]

示例：
    python generate-fund-evaluation.py 519702
    python generate-fund-evaluation.py 519702 --template ~/Desktop/基金评价底稿.xlsx --output-dir ~/Reports
    python generate-fund-evaluation.py 519702 --fetch-online  # 强制从天天基金补全数据

数据源优先级：
    1. 本地 JSON/Excel/CSV 数据
    2. 天天基金网 (fund.eastmoney.com) 在线爬取
    3. Wind/Choice 终端数据（如本地有）
"""

import argparse
import os
import sys
import glob
import json
import math
import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# 添加脚本所在目录到路径，以便导入 eastmoney_fetcher
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)


def find_local_data(fund_code, work_dir):
    """自动搜索与基金代码相关的本地数据文件"""
    patterns = [
        f"*{fund_code}*_holdings.csv",
        f"*fund_{fund_code}_complete_data.json",
        f"*最新风险收益指标*({fund_code}*).xlsx",
        f"*{fund_code}*_投资研究备忘录.md",
        f"*fund_diagnosis*_{fund_code}*.html",
        f"*基金诊断报告*_{fund_code}*.html",
    ]
    found = {}
    for pat in patterns:
        matches = glob.glob(os.path.join(work_dir, pat))
        if matches:
            key = pat.replace("*", "").replace(fund_code, "")
            found[key] = sorted(matches)[-1]
    return found


def load_holdings(fund_code, work_dir):
    """加载各季度持仓 CSV"""
    files = sorted(glob.glob(os.path.join(work_dir, f"{fund_code}_*Q_holdings.csv")))
    holdings = {}
    for f in files:
        basename = os.path.basename(f)
        quarter = basename.replace(f"{fund_code}_", "").replace("_holdings.csv", "")
        try:
            df = pd.read_csv(f, encoding="utf-8-sig", thousands=",")
            df["占净值比数值"] = df["占净值比"].astype(str).str.replace("%", "").astype(float)
            holdings[quarter] = df
        except Exception as e:
            print(f"  ⚠️ 读取持仓文件失败 {f}: {e}")
    return holdings


def load_json_data(fund_code, work_dir):
    """加载 fund_{fundcode}_complete_data.json"""
    path = os.path.join(work_dir, f"fund_{fund_code}_complete_data.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_risk_excel(fund_code, work_dir):
    """加载 最新风险收益指标 Excel"""
    matches = glob.glob(os.path.join(work_dir, f"*最新风险收益指标*({fund_code}*).xlsx"))
    if not matches:
        return pd.DataFrame()
    df = pd.read_excel(matches[-1], sheet_name=0, header=None)
    return df


def extract_risk_metrics_from_excel(risk_df):
    """从风险收益指标 Excel 中提取关键数据"""
    metrics = {}
    if risk_df.empty or risk_df.shape[1] < 3:
        return metrics
    
    for i in range(len(risk_df)):
        row = risk_df.iloc[i]
        vals = [str(v) if pd.notna(v) else "" for v in row.values]
        if len(vals) >= 2:
            # 尝试匹配基金名称关键词，或直接用通用提取逻辑
            for j in range(i+1, min(i+8, len(risk_df))):
                r = risk_df.iloc[j]
                r_vals = [str(v) if pd.notna(v) else "" for v in r.values]
                key = r_vals[0].strip()
                val = r_vals[1].strip() if len(r_vals) > 1 else ""
                
                if "Sharpe" in key or "夏普" in key:
                    metrics["sharpe_1y"] = val
                elif "Calmar" in key or "卡玛" in key or "calmar" in key.lower():
                    metrics["calmar_1y"] = val
                elif "年化收益率" in key:
                    metrics["annual_return_1y"] = val
                elif "年化波动率" in key:
                    metrics["annual_vol_1y"] = val
                elif "最大回撤" in key:
                    metrics["max_drawdown_1y"] = val
                elif "日度正收益概率" in key:
                    metrics["daily_win_rate_1y"] = val
    return metrics


def fetch_online_data(fund_code):
    """从天天基金网获取在线数据"""
    try:
        from eastmoney_fetcher import EastMoneyFetcher, eastmoney_to_r9alpha
        fetcher = EastMoneyFetcher()
        em_data = fetcher.fetch_all(fund_code)
        r9data = eastmoney_to_r9alpha(em_data, fund_code)
        return r9data, em_data
    except ImportError as e:
        print(f"  ⚠️ 无法导入 eastmoney_fetcher 模块: {e}")
        return {}, {}
    except Exception as e:
        print(f"  ⚠️ 从天天基金获取数据失败: {e}")
        return {}, {}


def merge_data(data, online_data):
    """合并本地数据和在线数据（在线数据补充本地缺失项）"""
    filled_keys = []
    if not online_data:
        return data, filled_keys
    
    for key, val in online_data.items():
        local_val = data.get(key, "")
        if not local_val or local_val in ["", "待补充", "None", "None%", "None元 ()"]:
            if val and val not in ["", "None", "None%", "None元 ()"]:
                data[key] = val
                filled_keys.append(key)
    return data, filled_keys


def build_data_dict(fund_code, work_dir, fetch_online=False):
    """构建用于填充评价底稿的数据字典"""
    data = {
        "fund_code": fund_code,
        "fund_name": "",
        "manager": "",
        "manager_tenure": "",
        "awards": "",
        "fund_size": "",
        "established_date": "",
        "nav": "",
        "annual_return_ytd": "",
        "annual_return_1y": "",
        "annual_return_3y": "",
        "annual_return_5y": "",
        "annual_return_since": "",
        "return_2018": "",
        "return_2021": "",
        "return_2022": "",
        "return_2024": "",
        "return_2025": "",
        "rank_1y": "",
        "rank_3y": "",
        "rank_5y": "",
        "max_drawdown_3y": "",
        "annual_vol_3y": "",
        "sharpe_3y": "",
        "calmar_3y": "",
        "sortino_3y": "",
        "downside_std_3y": "",
        "tracking_error": "",
        "info_ratio": "",
        "excess_return": "",
        "excess_vol": "",
        "excess_sharpe": "",
        "recovery_days": "",
        "monthly_win_rate": "",
        "up_month_avg": "",
        "down_month_avg": "",
        "top10_ratio": "",
        "turnover": "",
        "inst_ratio": "",
        "inst_change": "",
        "top3_sectors": "",
        "pe_ttm": "",
        "brinson": "",
        "market_analysis": "",
        "market_outlook": "",
        "r9alpha_score": "",
        "r9alpha_comment": "",
    }
    
    # 1. JSON 数据
    json_data = load_json_data(fund_code, work_dir)
    if json_data:
        data["fund_name"] = json_data.get("fund_name_full", "")
        data["manager"] = json_data.get("manager", "")
        data["fund_size"] = f"{json_data.get('fund_size', '')}亿元 ({json_data.get('fund_size_date', '')})"
        data["established_date"] = json_data.get("established_date", "")
        data["nav"] = f"{json_data.get('nav', '')}元 ({json_data.get('nav_date', '')})"
        
        returns = json_data.get("returns", {})
        data["annual_return_ytd"] = f"{returns.get('ytd', '')}%" if returns.get('ytd') is not None else ""
        data["annual_return_1y"] = f"{returns.get('1y', '')}%" if returns.get('1y') is not None else ""
        data["annual_return_3y"] = f"{returns.get('3y', '')}%" if returns.get('3y') is not None else ""
        data["annual_return_5y"] = f"{returns.get('5y', '')}%" if returns.get('5y') is not None else ""
        data["annual_return_since"] = f"{returns.get('since_inception', '')}%" if returns.get('since_inception') is not None else ""
        
        rankings = json_data.get("rankings", {})
        data["rank_1y"] = rankings.get("1y", "")
        data["rank_3y"] = rankings.get("3y", "")
        data["rank_5y"] = rankings.get("5y", "")
        
        risk = json_data.get("risk_metrics", {})
        data["max_drawdown_3y"] = f"{risk.get('max_drawdown_3y', '')}%" if risk.get('max_drawdown_3y') is not None else ""
        data["annual_vol_3y"] = f"{risk.get('annualized_volatility_3y', '')}%" if risk.get('annualized_volatility_3y') is not None else ""
        data["sharpe_3y"] = risk.get("sharpe_ratio", "")
        data["calmar_3y"] = risk.get("calmar_ratio", "")
        data["recovery_days"] = f"{risk.get('recovery_days', '')}天" if risk.get('recovery_days') is not None else ""
        data["monthly_win_rate"] = f"{risk.get('positive_month_ratio', '')}%" if risk.get('positive_month_ratio') is not None else ""
        data["top10_ratio"] = f"{json_data.get('top10_ratio', '')}%" if json_data.get('top10_ratio') is not None else ""
        
        sectors = json_data.get("sector_allocation", [])
        if sectors:
            top3 = sorted(sectors, key=lambda x: x.get("ratio", 0), reverse=True)[:3]
            data["top3_sectors"] = "、".join([f"{s['name']}({s['ratio']}%)" for s in top3])
        
        data["awards"] = f"合计{sum([a['count'] for a in json_data.get('awards', [])])}次"
    
    # 2. 风险收益 Excel
    risk_df = load_risk_excel(fund_code, work_dir)
    risk_metrics = extract_risk_metrics_from_excel(risk_df)
    if risk_metrics:
        data["sharpe_1y"] = risk_metrics.get("sharpe_1y", "")
        data["calmar_1y"] = risk_metrics.get("calmar_1y", "")
        data["annual_return_1y_excel"] = risk_metrics.get("annual_return_1y", "")
        data["annual_vol_1y_excel"] = risk_metrics.get("annual_vol_1y", "")
        data["max_drawdown_1y_excel"] = risk_metrics.get("max_drawdown_1y", "")
        data["daily_win_rate_1y"] = risk_metrics.get("daily_win_rate_1y", "")
    
    # 3. 持仓数据
    holdings = load_holdings(fund_code, work_dir)
    if holdings:
        latest_q = sorted(holdings.keys())[-1]
        latest_df = holdings[latest_q]
        data["top10_ratio"] = f"{latest_df['占净值比数值'].sum():.2f}%"
        runfeng = latest_df[latest_df["股票名称"] == "润丰股份"]
        if not runfeng.empty:
            data["runfeng_ratio"] = f"{runfeng['占净值比数值'].values[0]:.2f}%"
    
    # 4. 天天基金在线数据补全
    online_data = {}
    em_raw = {}
    online_filled_keys = []
    if fetch_online:
        online_data, em_raw = fetch_online_data(fund_code)
        data, online_filled_keys = merge_data(data, online_data)
    
    return data, holdings, json_data, online_data, em_raw, online_filled_keys


def fill_excel_template(template_path, output_path, data):
    """填充 Excel 模板"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sheet1"]
    
    ws["A1"] = "基金评价底稿（偏股混合-主动）"
    ws["A2"] = f"基金名称：{data.get('fund_name', '')}    基金代码：{data.get('fund_code', '')}    基金分类：混合型-偏股"
    
    ws["H2"] = f"{data.get('fund_code', '')} 数据/分析"
    ws["H2"].font = Font(bold=True, size=11)
    ws["H2"].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    row_map = {
        3:  "manager",
        4:  "manager_tenure",
        5:  "awards",
        6:  "管理产品",
        7:  "fund_size",
        8:  "established_date",
        9:  "规模（份额）待补充",
        10: "fund_size",
        11: "top10_ratio",
        12: "turnover",
        13: "top10_ratio",
        14: "pe_ttm",
        15: "inst_ratio",
        16: "inst_change",
        17: "top3_sectors",
        18: "annual_return_ytd",
        19: "return_2024 / return_2025",
        20: "return_2018",
        21: "annual_return_since",
        22: "annual_return_3y",
        23: "max_drawdown_3y",
        24: "annual_vol_3y 方差",
        25: "annual_vol_3y",
        26: "downside_std_3y",
        27: "sharpe_3y",
        28: "calmar_3y",
        29: "sortino_3y",
        30: "excess_return",
        31: "excess_vol",
        32: "excess_maxdd",
        33: "excess_sharpe",
        34: "brinson",
        35: "跟踪偏离度",
        36: "tracking_error",
        37: "info_ratio",
        38: "max_drawdown_3y",
        39: "annual_vol_3y",
        40: "downside_std_3y",
        41: "recovery_days",
        42: "回撤恢复",
        43: "daily_win_rate_1y",
        44: "monthly_win_rate",
        45: "monthly_win_rate",
        46: "up_month_avg",
        47: "down_month_avg",
        48: "established_date",
        49: "公司管理规模待补充",
        50: "发行只均规模待补充",
        51: "单只≥100亿占比待补充",
        52: "份额变化率1待补充",
        53: "份额变化率2待补充",
        54: "基民损益待补充",
        55: "分额度限额占比待补充",
        56: "公司深度亏损占比待补充",
        57: "限额变化率待补充",
        58: "限额时点待补充",
        59: "研究员数量待补充",
        60: "团队考核待补充",
        61: "投决机制待补充",
        62: "研转投输送率待补充",
        63: "考核机制待补充",
        64: "market_analysis",
        65: "market_outlook",
        66: "r9alpha_comment",
    }
    
    for row_idx, key in row_map.items():
        val = data.get(key, "")
        if not val or val == "None" or val == "None元 ()" or val == "None%":
            val = "待补充"
        ws.cell(row=row_idx+1, column=8, value=str(val))
        ws.cell(row=row_idx+1, column=8).alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.column_dimensions['H'].width = 70
    wb.save(output_path)
    print(f"  ✅ Excel 底稿已保存: {output_path}")


def generate_markdown_report(fund_code, output_path, data, holdings, json_data, online_data=None, em_raw=None, online_filled_keys=None):
    """生成 Markdown 深度评价报告"""
    
    online_filled_keys = online_filled_keys or []
    def src(key):
        return '天天基金' if key in online_filled_keys else '本地'
    
    # 提取在线持仓
    online_holdings = []
    if em_raw and 'holdings' in em_raw and not em_raw['holdings'].empty:
        for _, row in em_raw['holdings'].iterrows():
            online_holdings.append({
                'rank': row.get('rank', ''),
                'code': row.get('stock_code', ''),
                'name': row.get('stock_name', ''),
                'ratio': row.get('ratio', ''),
                'sector': row.get('sector', ''),
            })
    
    # 提取在线行业配置
    online_industry = []
    if em_raw and 'industry' in em_raw and not em_raw['industry'].empty:
        for _, row in em_raw['industry'].iterrows():
            online_industry.append({
                'name': row.get('industry', ''),
                'ratio': row.get('ratio', ''),
            })
    
    # 提取在线资产配置
    asset_str = ""
    if em_raw and 'asset_allocation' in em_raw and not em_raw['asset_allocation'].empty:
        latest = em_raw['asset_allocation'].iloc[0]
        parts = []
        for k, label in [('stock_ratio', '股票'), ('bond_ratio', '债券'), ('cash_ratio', '现金'), ('other_ratio', '其他')]:
            v = latest.get(k)
            if v is not None:
                parts.append(f"{label}: {v}%")
        asset_str = " | ".join(parts)
    
    lines = [
        f"# 基金评价报告：{data.get('fund_name', fund_code)}（{fund_code}）",
        "## 按照 R9Alpha 研基标准评价体系",
        "",
        "---",
        "",
        "## 📋 执行摘要",
        "",
        f"**基金名称**：{data.get('fund_name', '')}",
        f"**基金代码**：{fund_code}",
        f"**基金经理**：{data.get('manager', '')}",
        f"**基金规模**：{data.get('fund_size', '')}",
        f"**成立日期**：{data.get('established_date', '')}",
        f"**最新净值**：{data.get('nav', '')}",
        f"**比较基准**：{data.get('benchmark', '')}",
        "",
        "---",
        "",
        "## 一、基础信息分析",
        "",
        f"### 1.1 基金经理\n\n- 基金经理：{data.get('manager', '')}\n- 任职时长：{data.get('manager_tenure', '')}\n- 管理规模：{data.get('fund_size', '')}\n- 获奖记录：{data.get('awards', '')}",
        "",
        "### 1.2 收益指标\n",
        "| 区间 | 收益 | 排名 |",
        "|------|------|------|",
        f"| 近一周 | {data.get('return_1w', '')} | -- |",
        f"| 近一月 | {data.get('return_1m', '')} | -- |",
        f"| 近三月 | {data.get('return_3m', '')} | -- |",
        f"| 近六月 | {data.get('return_6m', '')} | -- |",
        f"| 近一年 | {data.get('annual_return_1y', '')} | {data.get('rank_1y', '')} |",
        f"| 近三年 | {data.get('annual_return_3y', '')} | {data.get('rank_3y', '')} |",
        f"| 近五年 | {data.get('annual_return_5y', '')} | {data.get('rank_5y', '')} |",
        f"| 成立以来 | {data.get('annual_return_since', '')} | -- |",
        "",
        "### 1.3 风险指标\n",
        "| 指标 | 数据 | 来源 |",
        "|------|------|------|",
        f"| 最大回撤 | {data.get('max_drawdown_3y', '')} | {src('max_drawdown_3y')} |",
        f"| 年化波动率 | {data.get('annual_vol_3y', '')} | {src('annual_vol_3y')} |",
        f"| 夏普比率 | {data.get('sharpe_3y', '')} | {src('sharpe_3y')} |",
        f"| 卡玛比率 | {data.get('calmar_3y', '')} | {src('calmar_3y')} |",
        f"| 索提诺比率 | {data.get('sortino_3y', '')} | {src('sortino_3y')} |",
        f"| 下行标准差 | {data.get('downside_std_3y', '')} | {src('downside_std_3y')} |",
        f"| 回撤恢复天数 | {data.get('recovery_days', '')} | {src('recovery_days')} |",
        f"| 月度胜率 | {data.get('monthly_win_rate', '')} | {src('monthly_win_rate')} |",
        f"| 上涨月平均收益 | {data.get('up_month_avg', '')} | {src('up_month_avg')} |",
        f"| 下跌月平均亏损 | {data.get('down_month_avg', '')} | {src('down_month_avg')} |",
        "",
        "### 1.4 持仓结构\n",
        f"- 前十大占比：{data.get('top10_ratio', '')}",
        f"- 前三大行业：{data.get('top3_sectors', '')}",
    ]
    
    # 在线持仓明细
    if online_holdings:
        lines.extend([
            "",
            "### 1.5 十大重仓股（天天基金最新）\n",
            "| 排名 | 代码 | 名称 | 占比 | 行业 |",
            "|------|------|------|------|------|",
        ])
        for h in online_holdings[:10]:
            lines.append(f"| {h['rank']} | {h['code']} | {h['name']} | {h['ratio']}% | {h['sector']} |")
    
    # 在线行业配置
    if online_industry:
        lines.extend([
            "",
            "### 1.6 行业配置（天天基金）\n",
            "| 行业 | 占比 |",
            "|------|------|",
        ])
        for ind in online_industry[:8]:
            lines.append(f"| {ind['name']} | {ind['ratio']}% |")
    
    # 资产配置
    if asset_str:
        lines.extend([
            "",
            f"### 1.7 资产配置：{asset_str}",
        ])
    
    # 持有人结构
    if data.get('inst_ratio'):
        lines.extend([
            "",
            f"### 1.8 持有人结构：机构占比 {data.get('inst_ratio', '')}",
        ])
    
    lines.extend([
        "",
        "## 二、R9Alpha 研基综合评价",
        "",
        f"{data.get('r9alpha_comment', '待补充')}",
        "",
        "---",
        "",
        "**数据来源**：本地持仓数据、天天基金网(fund.eastmoney.com)、Wind/Choice数据、基金季报",
        "**免责声明**：本报告仅供参考，不构成投资建议。基金投资有风险，入市需谨慎。",
    ])
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  ✅ Markdown 报告已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="R9Alpha 研基基金评价报告生成器 V2")
    parser.add_argument("fund_code", help="基金代码，如 519702")
    parser.add_argument("--template", default="", help="R9alpha 评价底稿模板 Excel 路径")
    parser.add_argument("--output-dir", default=os.getcwd(), help="输出目录")
    parser.add_argument("--work-dir", default=os.getcwd(), help="工作目录（搜索本地数据）")
    parser.add_argument("--fetch-online", action="store_true", 
                        help="强制从天天基金网(fund.eastmoney.com)在线爬取数据补全缺失指标")
    args = parser.parse_args()
    
    fund_code = args.fund_code
    work_dir = args.work_dir
    output_dir = args.output_dir
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"\n{'='*60}")
    print(f" 开始生成 {fund_code} 的 R9Alpha 评价报告")
    print(f"{'='*60}")
    
    # 搜索本地数据
    local_files = find_local_data(fund_code, work_dir)
    if local_files:
        print("\n📁 发现的本地数据文件：")
        for k, v in local_files.items():
            print(f"  - {k}: {v}")
    else:
        print("\n⚠️ 未在工作目录发现相关本地数据文件")
    
    # 判断是否启用在线获取
    fetch_online = args.fetch_online
    if not local_files and not fetch_online:
        print("\n⚠️ 本地无数据，自动启用天天基金在线数据获取...")
        fetch_online = True
    
    # 构建数据字典（含在线补全）
    data, holdings, json_data, online_data, em_raw, online_filled_keys = build_data_dict(fund_code, work_dir, fetch_online=fetch_online)
    
    # 生成 Markdown
    md_path = os.path.join(output_dir, f"基金评价报告_{fund_code}_R9Alpha.md")
    generate_markdown_report(fund_code, md_path, data, holdings, json_data, online_data, em_raw, online_filled_keys)
    
    # 生成 Excel（如果有模板）
    if args.template and os.path.exists(args.template):
        excel_path = os.path.join(output_dir, f"基金评价底稿_{fund_code}_R9Alpha.xlsx")
        fill_excel_template(args.template, excel_path, data)
    else:
        print(f"\n⚠️ 未提供模板或模板不存在，跳过 Excel 生成")
        print(f"   如需生成 Excel，请使用 --template 参数指定模板路径")
    
    # 汇总数据完整度
    empty_count = sum(1 for v in data.values() if not v or v in ["", "待补充", "None", "None%", "None元 ()"])
    total_count = len(data)
    filled_count = total_count - empty_count
    
    print(f"\n{'='*60}")
    print(f" 生成完成！")
    print(f" 数据完整度: {filled_count}/{total_count} ({filled_count/total_count*100:.1f}%)")
    if fetch_online and online_filled_keys:
        print(f" 天天基金补全: {len(online_filled_keys)} 个字段")
        print(f"   {', '.join(online_filled_keys[:10])}{'...' if len(online_filled_keys) > 10 else ''}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
